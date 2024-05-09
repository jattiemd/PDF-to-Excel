import os
import threading
import pandas as pd
import tabula
import pythoncom
from flask import flash, redirect, request, session
from win32com.client.gencache import EnsureDispatch
from openpyxl import load_workbook


UPLOAD_FOLDER = 'file_handler\\'
lock = threading.Lock()


def do_conversion(file):
    """
    Takes in pdf file, runs check if it has tables on or not. 
    If tables exist excel file will be created and saved for further processing.
    """

    if not file.filename.endswith('.pdf'):
        flash("File extension must be '.pdf' only.")
        threading.Timer(1.0, run_file_check_on, args=(session.get('PDF_FILE_NAME'),))
        threading.Timer(1.0, run_file_check_on, args=(session.get('EXCEL_FILE_NAME'),))
        threading.Timer(1.0, run_file_check_on, args=(session.get('NEW_EXCEL_FILE_NAME'),))

        return redirect(request.url)
    
    run_file_check_on(session.get('PDF_FILE_NAME', None))

    session['PDF_FILE_NAME'] = file.filename
    session['UPLOADS'].append(session['PDF_FILE_NAME'])
    file.save(os.path.join(UPLOAD_FOLDER, session['PDF_FILE_NAME']))
    
    print(f"* File '{session['PDF_FILE_NAME']}' has been saved!")
    print('* Converting...')

    lock.acquire()
    try:
        # Excel conversion
        # Getting Files
        extracted_file_path = os.path.join(UPLOAD_FOLDER, session['PDF_FILE_NAME'])
        pdf_file = extracted_file_path
        excel_file = extracted_file_path.replace('.pdf', '.xlsx')

        # Reading tables
        tables = tabula.read_pdf(pdf_file, pages='all', guess=False, stream=True)
        
        # Writing tables to excel file
        if tables:        
            print('* Tables exist: True')
            excel_writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
            for i, table in enumerate(tables):
                # Removing common recurring cells within tables
                common_cells = ["C:", "Printed", "Page"]
                for index, row in table.iterrows():
                    for column in table.columns:
                        if isinstance(row[column], str):
                            for common_cell in common_cells:
                                if row[column].startswith(common_cell):
                                    table.at[index, column] = pd.NA

                table.to_excel(excel_writer, sheet_name=f'Sheet_{i+1}', index=False)

            excel_writer.close()

            run_file_check_on(session.get('EXCEL_FILE_NAME', None))
            
            session['EXCEL_FILE_NAME'] = excel_file.replace(UPLOAD_FOLDER, '')
            session['UPLOADS'].append(session['EXCEL_FILE_NAME'])
            print('* Conversion Successful!')
            print(f"* File '{session["EXCEL_FILE_NAME"]}' contents displayed successfully!")
            flash('Conversion completed!')
            flash("Use the check boxes to select the tables that you want to generate for your Excel file. Then click 'Generate Excel' at the bottom of the page ")
        else:
            print('* Tables exist: False')
            run_file_check_on(session.get('EXCEL_FILE_NAME', None))
    finally:
        lock.release()

    redirect(request.url)


def remove_files(pdf_file_name, excel_file_name, new_excel_file_name, zip_folder, password_file):
    """Deleting files from file handler dir. Only if files exist within the dir"""

    lock.acquire()
    try: 
        print('* Locking file resources')
        print('* Rinsing file handler')
        # PDF
        pdf = os.path.join(UPLOAD_FOLDER, pdf_file_name)
        if pdf_file_name == None:
            print('* No Pdf file found')
        elif os.path.exists(pdf):
            os.remove(pdf)
            print('* Pdf file found')
            print('* Removing pdf file...')

        # EXCEL
        excel = os.path.join(UPLOAD_FOLDER, excel_file_name)
        if excel_file_name == None:
            print('* No Excel file found')
        elif os.path.exists(excel):
            os.remove(excel)
            print('* Excel file found')
            print('* Removing Excel file...')

        # NEW EXCEL
        new_excel = os.path.join(UPLOAD_FOLDER, new_excel_file_name)
        if new_excel_file_name == None:
            print('* No New Excel file found')
        elif os.path.exists(new_excel):
            os.remove(new_excel)
            print('* New Excel file found')
            print('* Removing New Excel file...')

        # ZIP FOLDER
        zip_folder = os.path.join(UPLOAD_FOLDER, zip_folder)
        if zip_folder == None:
            print('* No Zip folder found')
        elif os.path.exists(zip_folder):
            os.remove(zip_folder)
            print('* Zip folder found')
            print('* Removing Zip folder...')

        # PASSWORD FILE
        password_file = os.path.join(UPLOAD_FOLDER, password_file)
        if password_file == None:
            print('* No Password file found')
        elif os.path.exists(password_file):
            os.remove(password_file)
            print('* Password file found')
            print('* Removing Password file...')
    finally:
        lock.release()


def run_file_check_on(file_name):
    """Check if a file already exists. Remove if it if does"""

    if file_name is not None:
        os.remove(os.path.join(UPLOAD_FOLDER, file_name))


def password_protect_excel(file_dir_path, password):
    """Password protect entire excel workbook"""
    
    lock.acquire()
    try:
        pythoncom.CoInitialize()
        xl_file = EnsureDispatch("Excel.Application")
        wb = xl_file.Workbooks.Open(file_dir_path)
        xl_file.DisplayAlerts = False
        wb.Visible = False
        wb.SaveAs(file_dir_path, Password=password)
        wb.Close()
        xl_file.Quit()
        print("* File protected successfully")
        pythoncom.CoUninitialize()
    finally:
        lock.release()


def password_protect_sheets(file_dir_path, password):
    """Passsword protect excel worksheets only"""

    lock.acquire()
    try:
        wb = load_workbook(filename=file_dir_path)
        sheets = wb.sheetnames
        for sheet in sheets:
            wb[sheet].protection.set_password(password)
        wb.save(file_dir_path)
        print(f"* {len(wb.sheetnames)} sheet(s) protected successfully")
    finally:
        lock.release()
    